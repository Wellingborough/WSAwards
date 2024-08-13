import openpyxl
import sqlite3
import constants
import smtplib
from email.mime.text import MIMEText

print("Starting...")
print("Opening Database...")
connection = sqlite3.connect(constants.DATABASE_NAME)
cursor = connection.cursor()
cursor.execute(constants.CREATE_PUPIL_TABLE)

departmentNames = []
newPupilRecords = 0

#
# Use the contacts file to build and check the pupil database
#
wbContactData = None
try:
    wbContactData = openpyxl.load_workbook(constants.CONTACTS_SPREADSHEET)
except FileNotFoundError:
    print(f"Could not open the Contacts file: {constants.CONTACTS_SPREADSHEET}")
    exit(0)

# print(f"Opened the Contacts file: {constants.CONTACTS_SPREADSHEET}")

for sheet in wbContactData:
    if sheet.title == "Export":
        pass
        # print("Found the main sheet")

    foundHeaders = False

    for row in sheet.iter_rows(values_only=True):
        if row[constants.CONTACT_PREFERRED_NAME] == "Preferred Name":
            # print("Found headers")
            foundHeaders = True
            continue
        elif row[constants.CONTACT_PREFERRED_NAME] is None and foundHeaders:
            # print("Found end marker")
            break
        elif row[constants.CONTACT_PREFERRED_NAME] is not None:
            #
            # Pupil Email Address is of the format nnnnnn@wellingboroughschool.org
            #
            preferredForename = row[constants.CONTACT_PREFERRED_NAME].strip()
            surname = row[constants.CONTACT_SURNAME].strip()
            academicYear = row[constants.CONTACT_YEAR]
            email = row[constants.CONTACT_EMAIL].strip()

            emailSplit = email.split('@')
            puid = emailSplit[0]
            # print(f"Pupil: {puid}, Year: {academicYear}, Name: {preferredForename, surname}")

            #
            # Check whether this pupil exists in the database
            #
            cursor.execute(constants.CHECK_PUPIL_TABLE + puid)
            pupilCountResult = cursor.fetchone()
            pupilCount = pupilCountResult[0]

            #
            # If the pupil does not exist, create their record...
            #
            if pupilCount == 0:
                # print(f"Creating a new record for {puid}, {preferredForename, surname}")
                cursor.execute(constants.INSERT_PUPIL_ROW, (puid, academicYear, preferredForename, surname))
                newPupilRecords += 1
            else:
                # print(f"Checking the record for {puid}, {preferredForename, surname}")
                cursor.execute(constants.QUERY_PUPIL_TABLE + puid)
                pupilResult = cursor.fetchone()
                # print(pupilResult)


if newPupilRecords > 0:
    print(f"Created {str(newPupilRecords)} new pupil records")

wbContactData.close()
connection.commit()

#
# Now use the rewards spreadsheet to check for new awards, update the database count, and generate
# any postcards
#
wbMasterData = None
try:
    wbMasterData = openpyxl.load_workbook(constants.REWARDS_SPREADSHEET)
except FileNotFoundError:
    print(f"Could not open the Master file: {constants.REWARDS_SPREADSHEET}")
    exit(0)

# print(f"Opened the Master file: {constants.REWARDS_SPREADSHEET}")


for sheet in wbMasterData:
    departmentName = sheet.title

    if departmentName in constants.DEPARTMENT_EXCEPTIONS:
        if constants.DEPARTMENT_EXCEPTIONS[departmentName] == "":
            print(f"Skipping {departmentName}")
            continue
        else:
            print(f"Renaming {departmentName} to {constants.DEPARTMENT_EXCEPTIONS[departmentName]}")
            departmentName = constants.DEPARTMENT_EXCEPTIONS[departmentName]
    # print(f"Processing sheet {departmentName}")

    departmentNames.append(departmentName)

    if departmentName in constants.DEPARTMENT_HEADS:
        pass
        # print(f"Congratulations will come from {constants.DEPARTMENT_HEADS[departmentName]}")

    #
    # Ensure that the department table exists
    #
    sqlString = constants.CREATE_DEPT_TABLE.replace("{department}", departmentName)

    cursor.execute(sqlString)
    connection.commit()

    foundHeaders = False

    for row in sheet.iter_rows(values_only=True):
        if row[constants.PUPIL_NAME] == "Pupil Name":
            # print("Found headers")
            foundHeaders = True
            continue
        elif row[constants.PUPIL_NAME] is None and foundHeaders:
            # print("Found end marker")
            break
        elif row[constants.PUPIL_NAME] is not None:
            #
            # Points is a string of the form "x[y]"
            # Where x is the number of 'new' points and y is the cumulative total
            #
            pointsString = row[constants.PUPIL_POINTS]
            pointsList = pointsString.split('[')
            newPoints = int(pointsList[0])

            if newPoints == 0:
                continue

            # print(f"Pupil: {row[constants.PUPIL_NAME]}, Year: {row[constants.PUPIL_YEAR]}, Points: {newPoints}")

            #
            # Check whether this pupil exists in the database
            #
            fullname = row[constants.PUPIL_NAME]
            names = fullname.split(",")
            preferredForename = names[1].strip()
            surname = names[0].strip()
            academicYear = row[constants.PUPIL_YEAR]
            #
            # I cannot see a way, at present, to move this query into the constants file.
            # (because we need to intersperse values with the fixed parts)
            #

            #
            # Get PupilID from the pupil table
            #
            sqlString = constants.SELECT_PUPIL.replace("{fname}", preferredForename)
            sqlString = sqlString.replace("{sname}", surname)
            sqlString = sqlString.replace("{academicyear}", str(academicYear))

            cursor.execute(sqlString)
            pupilResult = cursor.fetchone()

            if pupilResult is None:
                print(f"Could not find pupil record for {preferredForename} {surname} in database")
                if academicYear < 7 or academicYear > 13:
                    print(f"Incorrect academic year ({str(academicYear)}) for Senior School" )

                continue
            pupilID = pupilResult[0]

            #
            # Now try to get the pupil's row from the department table
            #
            sqlString = constants.QUERY_DEPT_TABLE.replace("{department}", departmentName)
            cursor.execute(sqlString+str(pupilID))
            pupilResult = cursor.fetchone()
            oldPoints = 0

            if pupilResult is None:
                # print(f"Could not find pupil record for {departmentName} in database")
                sqlString = constants.INSERT_DEPT_ROW.replace("{department}", departmentName)
                # print(sqlString)
                cursor.execute(sqlString, (pupilID, oldPoints))
            else:
                oldPoints = pupilResult[0]

            if newPoints != oldPoints:
                sqlString = constants.UPDATE_DEPT_ROW.replace("{department}", departmentName)
                sqlString = sqlString.replace("{newpoints}", str(newPoints))

                # print("Updating database: "+str(oldPoints)+" versus "+str(newPoints))
                cursor.execute(sqlString+str(pupilID))



    # print(f"Finished {departmentName}")
wbMasterData.close()



print("Closing Database")
connection.commit()
connection.close()
print("Finished updating points totals")

#
# Initial attempt to generate report for a single pupil
#
print("Opening Database...")
connection = sqlite3.connect(constants.DATABASE_NAME)
cursor = connection.cursor()
innerCursor = connection.cursor()
pupilCount = 0
pupilWithAwardCount = 0


cursor.execute(constants.QUERY_ALL_PUPILS)
for row in cursor:
    pupilID = row[0]
    print("Pupil ID is", pupilID)
    pupilCount += 1

    pupilAwarded = False

    for departmentName in departmentNames:
        sqlString = constants.QUERY_DEPT_TABLE.replace("{department}", departmentName)
        innerCursor.execute(sqlString+str(pupilID))
        pointCount = innerCursor.fetchone()
        if pointCount is None:
            continue
        pointCount = pointCount[0]
        if not pupilAwarded:
            pupilAwarded = True
            pupilWithAwardCount += 1

        print(f"Total for {pupilID} in {departmentName} is {pointCount}")

print(f"Processed {pupilCount} Pupils, of whom {pupilWithAwardCount} had one or more awards")
print("Closing Database")
connection.commit()
connection.close()
print("Finished generating reports")



#
# Try out the email interface
#

# sender_email = "wellingboroughschoolrewards@gmail.com"
# sender_password = "nruk kaul bbju hssa"
# recipient_email = "christine.gamble@outlook.com"
# subject = "Congratulations from Wellingborough School"
# body = """
# <html>
#   <body>
#     <h1>Wellingborough School Rewards<h1>
#     <p><b>Congratulations</b>  You now been awarded a total of 57 points by the Physics Department</p>
#     <img src="https://www.wellingboroughschool.org/wp-content/uploads/2020/09/3-Science-432x282.jpg">
#     <p>Well done!<p>
#     <p>Mrs Hill<p>
#   </body>
# </html>
# """
# html_message = MIMEText(body, 'html')
# html_message['Subject'] = subject
# html_message['From'] = sender_email
# html_message['To'] = recipient_email
# with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
#    server.login(sender_email, sender_password)
#    server.sendmail(sender_email, recipient_email, html_message.as_string())